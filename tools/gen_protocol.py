#!/usr/bin/env python3
"""
R큐브 프로토콜 코드 생성기
==========================
docs/ 의 프로토콜 xlsx(단일 소스)를 파싱해 shared-protocol/ 아래에
C 헤더와 Python 모듈을 생성한다. 손으로 옮기지 않는다(로드맵 Phase 0 원칙).

파싱 대상:
  - STATEMENTS 시트  : OpCode 목록 (Callback Function ↔ Op Code)
  - APPENDIX 시트 F절 : NACK ResultCode 코드표

구조 상수(표준 헤더 오프셋 / TargetId / CAN 29비트 ID 레이아웃 /
우선순위 클래스 / ACK·NAK)는 CAN 시트 B·C절, APPENDIX A·E절의
고정 규약이므로 아래에 인코딩한다(출처 주석 표기).

사용:
    python tools/gen_protocol.py
    python tools/gen_protocol.py --check   # 생성 없이 파싱 결과만 출력
"""
from __future__ import annotations
import argparse
import glob
import os
import re
import sys
from datetime import date

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 필요: pip install openpyxl")

HEX2 = re.compile(r"^[0-9A-Fa-f]{2}$")
HEX_PREFIXED = re.compile(r"^0x[0-9A-Fa-f]{2}$")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(ROOT, "shared-protocol")


def find_xlsx() -> str:
    hits = glob.glob(os.path.join(ROOT, "docs", "*프로토콜*.xlsx"))
    if not hits:
        sys.exit("docs/ 에서 프로토콜 xlsx를 찾지 못했습니다.")
    return hits[0]


def find_extensions() -> list[str]:
    """확장 규격 문서(docs/R큐브_프로토콜_확장_*.md). 날짜 오름차순."""
    return sorted(glob.glob(os.path.join(ROOT, "docs", "*프로토콜_확장_*.md")))


MD_ROW = re.compile(r"^\|\s*([A-Za-z_][A-Za-z0-9_]*)\s*\|\s*([0-9A-Fa-f]{2})\s*\|"
                    r"\s*([^|]*)\|\s*([^|]*)\|")


def parse_extension_opcodes(path: str) -> list[dict]:
    """확장 문서의 <!-- GEN:OPCODES --> 표를 파싱한다.

    xlsx는 읽기전용 기준 문서라 손대지 않고, 이후 확정된 OpCode는 확장 문서에 적는다.
    생성기가 양쪽을 합쳐 헤더를 만들므로 '단일 소스' 원칙은 그대로 유지된다.
    """
    with open(path, encoding="utf-8") as f:
        text = f.read()
    m = re.search(r"<!--\s*GEN:OPCODES\s*-->(.*?)<!--\s*/GEN:OPCODES\s*-->", text, re.S)
    if not m:
        return []
    out = []
    for line in m.group(1).splitlines():
        row = MD_ROW.match(line.strip())
        if not row:
            continue
        name, code, category, ble_can = row.groups()
        if name.lower() == "이름" or set(code) <= {"-"}:
            continue
        out.append({
            "name": name.strip(),
            "code": int(code, 16),
            "category": category.strip(),
            "ble_can": ble_can.strip(),
            "src": os.path.basename(path),
        })
    return out


def merge_opcodes(base: list[dict], ext: list[dict]) -> list[dict]:
    """확장 OpCode를 병합한다. 이름/코드 중복은 즉시 실패시킨다(조용한 덮어쓰기 금지)."""
    by_name = {o["name"]: o for o in base}
    by_code = {o["code"]: o for o in base}
    for e in ext:
        if e["name"] in by_name:
            sys.exit(f"확장 OpCode 이름 중복: {e['name']} ({e['src']})")
        if e["code"] in by_code:
            dup = by_code[e["code"]]["name"]
            sys.exit(f"확장 OpCode 코드 중복: 0x{e['code']:02X} "
                     f"{e['name']} vs {dup} ({e['src']})")
        by_name[e["name"]] = e
        by_code[e["code"]] = e
        base.append(e)
    return base


def parse_opcodes(wb) -> list[dict]:
    ws = wb["STATEMENTS"]
    # 헤더 행(‘분류 | Callback Function | Op Code | 비고 | BLE/CAN’) 찾기
    hdr = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "분류" and ws.cell(r, 3).value == "Op Code":
            hdr = r
            break
    if hdr is None:
        sys.exit("STATEMENTS 헤더 행을 찾지 못했습니다.")
    out = []
    seen = set()
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(r, 2).value
        code = ws.cell(r, 3).value
        if not name or not code:
            continue
        code = str(code).strip()
        name = str(name).strip()
        if not HEX2.match(code):
            continue
        if name in seen:
            continue
        seen.add(name)
        out.append({
            "name": name,
            "code": int(code, 16),
            "category": str(ws.cell(r, 1).value or "").strip(),
            "ble_can": str(ws.cell(r, 5).value or "").strip(),
        })
    return out


def parse_resultcodes(wb) -> list[dict]:
    ws = wb["APPENDIX"]
    hdr = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "코드" and ws.cell(r, 2).value == "이름":
            hdr = r
            break
    if hdr is None:
        sys.exit("APPENDIX ResultCode 헤더(F절)를 찾지 못했습니다.")
    out = []
    for r in range(hdr + 1, ws.max_row + 1):
        code = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if code is None or name is None:
            continue
        code = str(code).strip()
        if not HEX_PREFIXED.match(code):
            # 표 끝(주석 행)에 도달
            if HEX2.match(code):
                pass
            else:
                continue
        out.append({
            "name": str(name).strip(),
            "code": int(code, 16),
            "meaning": str(ws.cell(r, 3).value or "").strip(),
        })
    return out


BANNER = """/*
 * {fname}
 * ----------------------------------------------------------------
 * ★ 자동 생성 파일 — 직접 수정 금지.
 *   원본: docs/{src}
 *   생성: tools/gen_protocol.py  ({today})
 *   수정이 필요하면 xlsx를 고치고 생성기를 다시 실행하세요.
 * ----------------------------------------------------------------
 */
"""

PY_BANNER = '''"""
{fname}
★ 자동 생성 파일 — 직접 수정 금지.
  원본: docs/{src}
  생성: tools/gen_protocol.py  ({today})
"""
'''


def gen_opcodes_h(ops, src) -> str:
    s = BANNER.format(fname="rcube_opcodes.h", src=src, today=date.today())
    s += "#pragma once\n#include <stdint.h>\n\n"
    s += f"/* 총 {len(ops)} 명령 (BLE/CAN 응용계층 공유) */\n"
    s += "typedef enum {\n"
    for o in ops:
        tag = f"  /* {o['category']} */" if o["category"] else ""
        s += f"    RCUBE_OP_{o['name']:<26} = 0x{o['code']:02X},{tag}\n"
    s += "} rcube_opcode_t;\n"
    return s


def gen_resultcodes_h(rcs, src) -> str:
    s = BANNER.format(fname="rcube_resultcodes.h", src=src, today=date.today())
    s += "#pragma once\n#include <stdint.h>\n\n"
    s += "/* CmdAck(0xAF) ResultCode — ACK 시 0x00(OK) */\n"
    s += "typedef enum {\n"
    for rc in rcs:
        s += f"    RCUBE_RC_{rc['name']:<16} = 0x{rc['code']:02X},  /* {rc['meaning']} */\n"
    s += "} rcube_result_t;\n"
    return s


PROTOCOL_H_BODY = """#pragma once
#include <stdint.h>
#include "rcube_opcodes.h"
#include "rcube_resultcodes.h"

/* =====================================================================
 * 표준 패킷 헤더 (APPENDIX A/D · CAN 시트 A절)
 *   [0] TargetId  [1] OpCode  [2:3] PacketSize(uint16, 와이어에서 BE)
 *   [4...] Property + Value (= CAN 데이터필드와 동일)
 * ===================================================================== */
typedef struct __attribute__((packed)) {
    uint8_t  target_id;    /* [0]   대상 주소 (아래 RCUBE_ADDR_*) */
    uint8_t  op_code;      /* [1]   rcube_opcode_t */
    uint16_t packet_size;  /* [2:3] 전체 패킷 길이 — 와이어에서 big-endian */
} rcube_header_t;

/* ---- TargetId / 주소 규약 (APPENDIX A) ---- */
#define RCUBE_ADDR_NODE_MIN   0x01u   /* 개별 노드 0x01~0x08 */
#define RCUBE_ADDR_NODE_MAX   0x08u
#define RCUBE_ADDR_HUB        0xFEu   /* 대표(edge central=노드01) */
#define RCUBE_ADDR_BROADCAST  0xFFu   /* 전 큐브 브로드캐스트 */

/* ---- CmdAck Status (APPENDIX E, ASCII ACK/NAK) ---- */
#define RCUBE_ACK  0x06u
#define RCUBE_NAK  0x15u

/* =====================================================================
 * CAN 29비트 확장 ID 레이아웃 (CAN 시트 B절)
 *   [28:26] Priority | [25:18] OpCode | [17] MULTI | [16] FLAG
 *   [15:8]  SrcId    | [7:0]   DstId
 * ===================================================================== */
#define RCUBE_CAN_ID(pri, op, multi, flag, src, dst)  \\
    ( ((uint32_t)((pri)  & 0x7u ) << 26) |            \\
      ((uint32_t)((op)   & 0xFFu) << 18) |            \\
      ((uint32_t)((multi)& 0x1u ) << 17) |            \\
      ((uint32_t)((flag) & 0x1u ) << 16) |            \\
      ((uint32_t)((src)  & 0xFFu) <<  8) |            \\
      ((uint32_t)((dst)  & 0xFFu)      ) )

#define RCUBE_CAN_PRI(id)    (((id) >> 26) & 0x7u)
#define RCUBE_CAN_OPCODE(id) (((id) >> 18) & 0xFFu)
#define RCUBE_CAN_MULTI(id)  (((id) >> 17) & 0x1u)
#define RCUBE_CAN_FLAG(id)   (((id) >> 16) & 0x1u)
#define RCUBE_CAN_SRC(id)    (((id) >>  8) & 0xFFu)
#define RCUBE_CAN_DST(id)    ( (id)        & 0xFFu)

#define RCUBE_CAN_SRC_MASTER 0xFEu   /* PC 또는 edge central */

/* =====================================================================
 * CAN 멀티프레임 (MULTI=1) — 확장 규격 §5
 *   MULTI=0 : 단일 프레임(데이터필드 전체가 페이로드)
 *   MULTI=1 : 데이터필드 [0]=세그먼트 헤더, [1:7]=페이로드 조각
 *       세그먼트 헤더  bit7=FIRST, bit6=LAST, bit5:0=순번(0~63)
 *       FIRST 세그먼트의 페이로드 앞 2바이트 = 전체 길이(BE16)
 *       → FIRST 데이터 5바이트, 이후 7바이트씩. 최대 5+63*7 = 446바이트.
 *   재조립은 (SrcId, OpCode)별 버퍼 1개. 순번 불일치 또는 타임아웃이면 폐기한다.
 * ===================================================================== */
#define RCUBE_CAN_SEG_FIRST      0x80u
#define RCUBE_CAN_SEG_LAST       0x40u
#define RCUBE_CAN_SEG_INDEX(h)   ((h) & 0x3Fu)
#define RCUBE_CAN_SEG_MAX_INDEX  0x3Fu
#define RCUBE_CAN_SEG_FIRST_DATA 5u      /* FIRST 세그먼트가 싣는 실데이터 바이트 */
#define RCUBE_CAN_SEG_DATA       7u      /* 이후 세그먼트가 싣는 바이트 */
#define RCUBE_CAN_REASSEMBLY_MAX 446u    /* 멀티프레임 최대 페이로드 */
#define RCUBE_CAN_REASSEMBLY_TIMEOUT_MS 200u

/* ---- CAN 우선순위 클래스 (CAN 시트 C절, 낮을수록 우선) ---- */
typedef enum {
    RCUBE_PRI_ESTOP        = 0,  /* D0 EmergencyStop 최우선 */
    RCUBE_PRI_SAFETY_SYNC  = 1,  /* D1·D2·D9·C7·B7 (B7=MotionComplete: 시퀀스 게이트) */
    RCUBE_PRI_MOTION       = 2,  /* C0~C3·C5·C8~CF */
    RCUBE_PRI_QUERY        = 3,  /* AF·B0~B6 */
    RCUBE_PRI_PERIPHERAL   = 4,  /* E0~E3·E5~E7·EA~ED */
    RCUBE_PRI_CONFIG       = 5,  /* D3~D8·DA·DB */
    RCUBE_PRI_MISSION_OTA  = 6,  /* F0~F8 */
    RCUBE_PRI_RESERVED     = 7,
} rcube_can_pri_t;
"""


def gen_protocol_h(src) -> str:
    return BANNER.format(fname="rcube_protocol.h", src=src, today=date.today()) + PROTOCOL_H_BODY


def gen_protocol_py(ops, rcs, src) -> str:
    s = PY_BANNER.format(fname="rcube_protocol.py", src=src, today=date.today())
    s += "from enum import IntEnum\n\n\n"
    s += "class OpCode(IntEnum):\n"
    for o in ops:
        s += f"    {o['name']} = 0x{o['code']:02X}\n"
    s += "\n\nclass ResultCode(IntEnum):\n"
    for rc in rcs:
        s += f"    {rc['name']} = 0x{rc['code']:02X}\n"
    s += "\n\n# ---- 주소 규약 (APPENDIX A) ----\n"
    s += "ADDR_NODE_MIN = 0x01\nADDR_NODE_MAX = 0x08\n"
    s += "ADDR_HUB = 0xFE       # 대표(edge central=노드01)\n"
    s += "ADDR_BROADCAST = 0xFF\n\n"
    s += "# ---- CmdAck Status ----\nACK = 0x06\nNAK = 0x15\n\n"
    s += "# ---- CAN 29비트 ID 레이아웃 (CAN 시트 B절) ----\n"
    s += "CAN_SRC_MASTER = 0xFE\n\n"
    s += "def can_id(pri, op, multi, flag, src, dst):\n"
    s += "    return ((pri & 0x7) << 26 | (op & 0xFF) << 18 | (multi & 1) << 17 |\n"
    s += "            (flag & 1) << 16 | (src & 0xFF) << 8 | (dst & 0xFF))\n\n"
    s += "class CanPri(IntEnum):\n"
    s += "    ESTOP = 0\n    SAFETY_SYNC = 1\n    MOTION = 2\n    QUERY = 3\n"
    s += "    PERIPHERAL = 4\n    CONFIG = 5\n    MISSION_OTA = 6\n    RESERVED = 7\n"
    return s


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--check", action="store_true", help="파싱 결과만 출력")
    args = ap.parse_args()

    src_path = find_xlsx()
    src = os.path.basename(src_path)
    wb = openpyxl.load_workbook(src_path, data_only=True)
    ops = parse_opcodes(wb)
    rcs = parse_resultcodes(wb)
    base_n = len(ops)

    ext_files = find_extensions()
    for p in ext_files:
        ext = parse_extension_opcodes(p)
        if ext:
            ops = merge_opcodes(ops, ext)
            print(f"[gen_protocol] 확장: docs/{os.path.basename(p)} "
                  f"→ OpCode {len(ext)}개 병합")
    if ext_files:
        src = src + " + " + ", ".join(os.path.basename(p) for p in ext_files)

    print(f"[gen_protocol] 원본: docs/{os.path.basename(src_path)}")
    print(f"[gen_protocol] OpCode {len(ops)}개(기준 {base_n} + 확장 {len(ops) - base_n}), "
          f"ResultCode {len(rcs)}개")

    if args.check:
        for o in ops:
            print(f"  OP 0x{o['code']:02X} {o['name']}")
        for rc in rcs:
            print(f"  RC 0x{rc['code']:02X} {rc['name']}")
        return 0

    os.makedirs(OUT_DIR, exist_ok=True)
    files = {
        "rcube_opcodes.h": gen_opcodes_h(ops, src),
        "rcube_resultcodes.h": gen_resultcodes_h(rcs, src),
        "rcube_protocol.h": gen_protocol_h(src),
        "rcube_protocol.py": gen_protocol_py(ops, rcs, src),
    }
    for name, content in files.items():
        path = os.path.join(OUT_DIR, name)
        with open(path, "w", encoding="utf-8") as fp:
            fp.write(content)
        print(f"  생성: shared-protocol/{name}  ({len(content)} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
